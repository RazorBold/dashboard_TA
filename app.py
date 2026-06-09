from flask import Flask, render_template, request, jsonify, send_file
from flask_socketio import SocketIO, emit
from flask_cors import CORS
from database import init_db, get_db
from dotenv import load_dotenv
import io
import json
import datetime
import time
import random
import logging
import os

load_dotenv()

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

app = Flask(__name__)
app.config["SECRET_KEY"] = "ble-monitor-secret-2024"
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="eventlet")

# ─── Config ───────────────────────────────────────────────────────────────────

MQTT_HOST    = os.getenv("MQTT_HOST", "36.92.47.218")
MQTT_PORT    = int(os.getenv("MQTT_PORT", "14583"))
MQTT_ENABLED = os.getenv("MQTT_ENABLED", "true").lower() == "true"

TOPIC_EVENTS    = "ble/events/#"
TOPIC_HEARTBEAT = "ble/heartbeat/#"

# ─── Detection Buffer (RSSI Voting) ──────────────────────────────────────────
# Tampung deteksi BLE sebelum diproses.
# Setelah BULK_MIN_COUNT deteksi (total semua gateway) ATAU BULK_WINDOW_SECS detik,
# pilih gateway dengan rata-rata RSSI terbaik → proses 1 event.
BULK_MIN_COUNT   = int(os.getenv("BULK_MIN_COUNT",  "4"))   # jumlah deteksi minimum
BULK_WINDOW_SECS = float(os.getenv("BULK_WINDOW_SECS", "8"))   # max detik tunggu

# RSSI threshold: cegah false-positive saat container jauh dari gate tapi masih kedeteksi
# OUT lebih ketat: hanya catat jika container benar-benar dekat gate keluar
# IN lebih longgar: -99 = tidak ada filter (sinyal lemah dari gate masuk masih valid)
RSSI_MIN_OUT = int(os.getenv("RSSI_MIN_OUT", "-65"))   # misal -65 dBm = jarak < ±5m
RSSI_MIN_IN  = int(os.getenv("RSSI_MIN_IN",  "-99"))   # -99 = no filter untuk IN

# { mac: { "gw": {"GW-001": {"n": int, "sum": int}}, "first": float, "total": int } }
_det_buf: dict = {}


def _buf_add(mac: str, gateway_id: str, rssi: int) -> list:
    """
    Tambah 1 deteksi ke buffer. Return list [(mac, gw, avg_rssi)] jika sudah
    waktunya flush (count terpenuhi atau window habis), else list kosong.
    """
    now = time.time()
    if mac not in _det_buf:
        _det_buf[mac] = {"gw": {}, "first": now, "total": 0}
    entry = _det_buf[mac]
    if gateway_id not in entry["gw"]:
        entry["gw"][gateway_id] = {"n": 0, "sum": 0}
    entry["gw"][gateway_id]["n"]   += 1
    entry["gw"][gateway_id]["sum"] += rssi
    entry["total"] += 1

    age = now - entry["first"]
    if entry["total"] >= BULK_MIN_COUNT or age >= BULK_WINDOW_SECS:
        return [_buf_flush(mac)]
    return []


def _buf_flush(mac: str) -> tuple:
    """
    Ambil keputusan dari buffer: gateway dengan rata-rata RSSI tertinggi.
    Hapus entry dari buffer. Return (mac, best_gateway_id, avg_rssi).
    """
    entry = _det_buf.pop(mac, None)
    if not entry or not entry["gw"]:
        return (mac, None, -70)
    best_gw, stats = max(
        entry["gw"].items(),
        key=lambda kv: kv[1]["sum"] / kv[1]["n"]
    )
    avg_rssi = int(stats["sum"] / stats["n"])
    detail = " | ".join(
        f"{gw}:{v['n']}x avg:{int(v['sum']/v['n'])} dBm"
        for gw, v in entry["gw"].items()
    )
    log.info(f"BUFFER FLUSH ({entry['total']} deteksi, {entry['gw'].__len__()} gw) "
             f"→ pemenang: {best_gw} avg {avg_rssi} dBm | [{detail}]")
    return (mac, best_gw, avg_rssi)


def _buf_flush_stale() -> list:
    """Flush semua entry yang sudah melewati BULK_WINDOW_SECS. Dipanggil periodik."""
    now   = time.time()
    stale = [m for m, e in _det_buf.items() if now - e["first"] >= BULK_WINDOW_SECS]
    return [_buf_flush(m) for m in stale]


# ─── Core BLE Logic (shared by HTTP + MQTT) ───────────────────────────────────

def process_ble_event(mac: str, gateway_id: str, rssi: int = -70) -> dict:
    """
    Proses deteksi BLE tag menggunakan state-machine per container.
    Event hanya dicatat saat STATUS BERUBAH: OUT→IN atau IN→OUT.
    Skip tanpa error jika status sudah sama.
    """
    mac = mac.upper().strip()

    db = get_db()
    try:
        tag = db.execute(
            "SELECT * FROM ble_tags WHERE mac_address=? AND status='Active'", [mac]
        ).fetchone()
        if not tag:
            return {"ok": False, "reason": "tag not found"}

        gw = db.execute(
            "SELECT * FROM gateways WHERE gateway_id=?", [gateway_id]
        ).fetchone()
        if not gw:
            return {"ok": False, "reason": "gateway not found"}

        event_type   = gw["zone"]          # IN atau OUT dari setting zone gateway
        container_id = tag["container_id"]

        # RSSI filter: tolak event jika sinyal terlalu lemah (container jauh dari gate)
        min_rssi = RSSI_MIN_OUT if event_type == "OUT" else RSSI_MIN_IN
        if rssi < min_rssi:
            log.info(f"RSSI WEAK: {container_id} {event_type} | {rssi} dBm < {min_rssi} dBm | gw={gateway_id} → skip")
            return {"ok": False, "reason": "rssi_weak", "rssi": rssi, "threshold": min_rssi}

        # State machine: cek status terakhir container di event_logs
        last = db.execute(
            "SELECT event_type FROM event_logs WHERE container_id=? ORDER BY timestamp DESC LIMIT 1",
            [container_id]
        ).fetchone()
        current_state = last["event_type"] if last else None

        if current_state == event_type:
            # Status sama → skip, tidak perlu dicatat
            log.info(f"STATE SAME: {container_id} sudah {event_type} | gw={gateway_id}")
            return {"ok": False, "reason": "same_state", "current": event_type}

        # Status berubah (None→IN, OUT→IN, IN→OUT) → catat event
        now_str = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        db.execute(
            "INSERT INTO event_logs (container_id, mac_address, gateway_id, event_type, timestamp, rssi) "
            "VALUES (?,?,?,?,?,?)",
            [container_id, mac, gateway_id, event_type, now_str, rssi]
        )
        db.execute("UPDATE ble_tags  SET last_seen=?, position=? WHERE mac_address=?", [now_str, event_type, mac])
        db.execute("UPDATE gateways  SET last_active=?, status='Online' WHERE gateway_id=?",
                   [now_str, gateway_id])
        db.commit()

        prev = current_state or "—"
        log.info(f"EVENT {prev}->{event_type} | {container_id} | {mac} | {gateway_id} | {rssi} dBm")
        event_row = {
            "container_id": container_id,
            "mac_address":  mac,
            "gateway_id":   gateway_id,
            "gateway_nama": gw["nama"],
            "event_type":   event_type,
            "timestamp":    now_str,
            "rssi":         rssi,
        }
        socketio.emit("new_event", event_row)
        return {"ok": True, "event": event_row}

    finally:
        db.close()


def process_heartbeat(gateway_id: str):
    """Update status gateway → Online saat terima heartbeat dari Pi."""
    db = get_db()
    db.execute(
        "UPDATE gateways SET status='Online', last_active=CURRENT_TIMESTAMP "
        "WHERE gateway_id=?", [gateway_id]
    )
    db.commit()
    db.close()
    log.debug(f"HEARTBEAT | {gateway_id}")


# ─── MQTT Subscriber (background task) ───────────────────────────────────────

def start_mqtt_client():
    """
    Jalankan paho-mqtt subscriber sebagai background task eventlet.
    Subscribe ke:
      ble/events/#     — deteksi BLE dari Pi
      ble/heartbeat/#  — heartbeat Pi
    """
    import eventlet
    try:
        import paho.mqtt.client as mqtt
    except ImportError:
        log.warning("paho-mqtt tidak terinstall. MQTT dinonaktifkan. Jalankan: pip install paho-mqtt")
        return

    def on_connect(client, userdata, flags, reason_code, properties):
        if reason_code.is_failure:
            log.error(f"MQTT connect failed: {reason_code}")
        else:
            client.subscribe(TOPIC_EVENTS)
            client.subscribe(TOPIC_HEARTBEAT)
            log.info(f"MQTT connected → {MQTT_HOST}:{MQTT_PORT} | subscribed to ble/#")

    def on_disconnect(client, userdata, flags, reason_code, properties):
        log.warning(f"MQTT disconnected ({reason_code}), akan reconnect...")

    def on_message(client, userdata, msg):
        topic = msg.topic
        try:
            payload = json.loads(msg.payload.decode("utf-8"))
        except Exception:
            log.warning(f"MQTT payload bukan JSON valid: {msg.payload}")
            return

        log.info(f"MQTT RX: {topic} | {payload}")

        if topic.startswith("ble/heartbeat/"):
            gw_id = payload.get("gateway_id") or topic.split("/")[-1]
            process_heartbeat(gw_id)
            return

        if topic.startswith("ble/events/"):
            mac        = payload.get("mac_address", "")
            gateway_id = payload.get("gateway_id", topic.split("/")[-1])
            rssi       = int(payload.get("rssi", -70))
            if mac:
                buf_n  = _det_buf.get(mac, {}).get("total", 0) + 1
                ts_now = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                log.info(f"BUFFER ADD : {mac} @ {gateway_id} ({rssi} dBm) [{buf_n}/{BULK_MIN_COUNT}]")
                socketio.emit("ble_raw", {
                    "type": "detect",
                    "mac": mac, "gw": gateway_id, "rssi": rssi, "ts": ts_now,
                    "n": buf_n, "max": BULK_MIN_COUNT
                })
                ready = _buf_add(mac, gateway_id, rssi)
                for m, gw, avg in ready:
                    if gw:
                        result = process_ble_event(m, gw, avg)
                        status = result.get("reason", "ok") if not result.get("ok") else "ok"
                        socketio.emit("ble_raw", {
                            "type": "flush", "status": status,
                            "mac": m, "gw": gw, "rssi": avg,
                            "ts": datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                            "n": BULK_MIN_COUNT, "max": BULK_MIN_COUNT,
                            "ev": result.get("event", {}).get("event_type", "")
                        })
                        if not result.get("ok"):
                            log.warning(f"process_ble_event SKIP: {status} | mac={m} gw={gw}")
            return

    client = mqtt.Client(
        callback_api_version=mqtt.CallbackAPIVersion.VERSION2,
        client_id="ble-monitor-server",
        clean_session=True
    )
    client.on_connect    = on_connect
    client.on_disconnect = on_disconnect
    client.on_message    = on_message
    client.reconnect_delay_set(min_delay=3, max_delay=30)

    try:
        client.connect(MQTT_HOST, MQTT_PORT, keepalive=60)
    except Exception as e:
        log.error(f"Tidak bisa connect ke MQTT broker {MQTT_HOST}:{MQTT_PORT} — {e}")
        log.info("Fallback: gunakan HTTP endpoint /api/ble/event")
        return

    # Jangan pakai loop_start() — thread paho konflik dengan eventlet greenlet.
    # Gunakan loop manual + yield ke eventlet hub setiap iterasi.
    log.info("MQTT loop manual dimulai (eventlet-safe)...")
    _last_stale_check = time.time()
    while True:
        client.loop(timeout=0.05)
        eventlet.sleep(0.05)

        # Flush buffer yang sudah timeout (1x per detik)
        now = time.time()
        if now - _last_stale_check >= 1.0:
            _last_stale_check = now
            for m, gw, avg in _buf_flush_stale():
                if gw:
                    log.info(f"BUFFER FLUSH (timeout): {m} -> {gw} avg {avg} dBm")
                    result = process_ble_event(m, gw, avg)
                    status = result.get("reason", "ok") if not result.get("ok") else "ok"
                    socketio.emit("ble_raw", {
                        "type": "flush", "status": status,
                        "mac": m, "gw": gw, "rssi": avg,
                        "ts": datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                        "n": BULK_MIN_COUNT, "max": BULK_MIN_COUNT,
                        "ev": result.get("event", {}).get("event_type", "")
                    })
                    if not result.get("ok"):
                        log.warning(f"process_ble_event SKIP: {status} | mac={m} gw={gw}")


# ─── Pages ────────────────────────────────────────────────────────────────────

@app.route("/")
def dashboard():
    return render_template("dashboard.html")

@app.route("/tags")
def tags_page():
    return render_template("tags.html")

@app.route("/gateways")
def gateways_page():
    return render_template("gateways.html")

@app.route("/logs")
def logs_page():
    return render_template("logs.html")

@app.route("/analytics")
def analytics_page():
    return render_template("analytics.html")

@app.route("/deteksi")
def deteksi_page():
    return render_template("deteksi.html")


# ─── API: Dashboard KPI ───────────────────────────────────────────────────────

@app.route("/api/dashboard/kpi")
def api_kpi():
    wilayah_id = request.args.get("wilayah_id")
    depo_id    = request.args.get("depo_id")
    db = get_db()

    gw_filter, gw_params = "", []
    if depo_id:
        gw_filter = "AND g.depo_id = ?";  gw_params = [depo_id]
    elif wilayah_id:
        gw_filter = "AND d.wilayah_id = ?"; gw_params = [wilayah_id]

    today = datetime.date.today().isoformat()

    def count(event):
        return db.execute(f"""
            SELECT COUNT(*) FROM event_logs e
            JOIN gateways g ON e.gateway_id = g.gateway_id
            JOIN depo d ON g.depo_id = d.id
            WHERE e.event_type=? AND date(e.timestamp)=? {gw_filter}
        """, [event, today] + gw_params).fetchone()[0]

    total_in  = count("IN")
    total_out = count("OUT")
    gw_online = db.execute(f"SELECT COUNT(*) FROM gateways g JOIN depo d ON g.depo_id=d.id WHERE g.status='Online' {gw_filter}", gw_params).fetchone()[0]
    gw_total  = db.execute(f"SELECT COUNT(*) FROM gateways g JOIN depo d ON g.depo_id=d.id WHERE 1=1 {gw_filter}", gw_params).fetchone()[0]
    tag_total = db.execute("SELECT COUNT(*) FROM ble_tags").fetchone()[0]
    db.close()

    return jsonify({
        "total_in": total_in, "total_out": total_out,
        "gw_online": gw_online, "gw_total": gw_total, "tag_total": tag_total
    })


@app.route("/api/dashboard/activity")
def api_activity():
    limit = int(request.args.get("limit", 20))
    db    = get_db()
    rows  = db.execute("""
        SELECT e.id, e.container_id, e.mac_address, e.gateway_id,
               e.event_type, e.timestamp, e.rssi,
               g.nama as gateway_nama, d.nama as depo_nama, w.nama as wilayah_nama
        FROM event_logs e
        LEFT JOIN gateways g ON e.gateway_id = g.gateway_id
        LEFT JOIN depo d ON g.depo_id = d.id
        LEFT JOIN wilayah w ON d.wilayah_id = w.id
        ORDER BY e.timestamp DESC LIMIT ?
    """, [limit]).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


# ─── API: Wilayah & Depo ──────────────────────────────────────────────────────

@app.route("/api/wilayah")
def api_wilayah():
    db = get_db()
    rows = db.execute("SELECT * FROM wilayah ORDER BY nama").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/depo")
def api_depo():
    wilayah_id = request.args.get("wilayah_id")
    db = get_db()
    if wilayah_id:
        rows = db.execute("SELECT * FROM depo WHERE wilayah_id=? ORDER BY nama", [wilayah_id]).fetchall()
    else:
        rows = db.execute("SELECT d.*, w.nama as wilayah_nama FROM depo d JOIN wilayah w ON d.wilayah_id=w.id ORDER BY d.nama").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


# ─── API: BLE Tags ────────────────────────────────────────────────────────────

@app.route("/api/tags", methods=["GET"])
def api_tags_list():
    db = get_db()
    rows = db.execute("""
        SELECT t.*, c.type as container_type
        FROM ble_tags t
        LEFT JOIN containers c ON t.container_id = c.container_id
        ORDER BY t.created_at DESC
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/tags", methods=["POST"])
def api_tags_create():
    data = request.json
    db = get_db()
    try:
        if data.get("container_id"):
            db.execute("INSERT OR IGNORE INTO containers (container_id, type) VALUES (?,?)",
                       [data["container_id"], data.get("container_type", "Standard")])
        db.execute("INSERT INTO ble_tags (mac_address, container_id, status) VALUES (?,?,?)",
                   [data["mac_address"], data.get("container_id"), data.get("status", "Active")])
        db.commit()
        db.close()
        return jsonify({"ok": True})
    except Exception as e:
        db.close()
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/tags/<int:tag_id>", methods=["PUT"])
def api_tags_update(tag_id):
    data = request.json
    db = get_db()
    try:
        if data.get("container_id"):
            db.execute("INSERT OR IGNORE INTO containers (container_id, type) VALUES (?,?)",
                       [data["container_id"], data.get("container_type", "Standard")])
        db.execute("UPDATE ble_tags SET mac_address=?, container_id=?, status=? WHERE id=?",
                   [data["mac_address"], data.get("container_id"), data.get("status", "Active"), tag_id])
        db.commit()
        db.close()
        return jsonify({"ok": True})
    except Exception as e:
        db.close()
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/tags/<int:tag_id>", methods=["DELETE"])
def api_tags_delete(tag_id):
    db = get_db()
    db.execute("DELETE FROM ble_tags WHERE id=?", [tag_id])
    db.commit()
    db.close()
    return jsonify({"ok": True})


# ─── API: Gateways ────────────────────────────────────────────────────────────

@app.route("/api/gateways", methods=["GET"])
def api_gateways_list():
    db = get_db()
    rows = db.execute("""
        SELECT g.*, d.nama as depo_nama, w.nama as wilayah_nama
        FROM gateways g
        LEFT JOIN depo d ON g.depo_id = d.id
        LEFT JOIN wilayah w ON d.wilayah_id = w.id
        ORDER BY g.created_at DESC
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/gateways", methods=["POST"])
def api_gateways_create():
    data = request.json
    db = get_db()
    try:
        db.execute("INSERT INTO gateways (gateway_id, nama, depo_id, zone, ip_address, status) VALUES (?,?,?,?,?,?)",
                   [data["gateway_id"], data["nama"], data.get("depo_id"),
                    data.get("zone", "IN"), data.get("ip_address"), data.get("status", "Offline")])
        db.commit()
        db.close()
        return jsonify({"ok": True})
    except Exception as e:
        db.close()
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/gateways/<int:gw_id>", methods=["PUT"])
def api_gateways_update(gw_id):
    data = request.json
    db = get_db()
    try:
        db.execute("UPDATE gateways SET gateway_id=?, nama=?, depo_id=?, zone=?, ip_address=?, status=? WHERE id=?",
                   [data["gateway_id"], data["nama"], data.get("depo_id"),
                    data.get("zone", "IN"), data.get("ip_address"), data.get("status", "Offline"), gw_id])
        db.commit()
        db.close()
        return jsonify({"ok": True})
    except Exception as e:
        db.close()
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/gateways/<int:gw_id>", methods=["DELETE"])
def api_gateways_delete(gw_id):
    db = get_db()
    db.execute("DELETE FROM gateways WHERE id=?", [gw_id])
    db.commit()
    db.close()
    return jsonify({"ok": True})


# ─── API: Event Logs ─────────────────────────────────────────────────────────

@app.route("/api/logs", methods=["GET"])
def api_logs():
    page     = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 25))
    search     = request.args.get("search", "")
    event_type = request.args.get("event_type", "")
    depo_id    = request.args.get("depo_id", "")
    wilayah_id = request.args.get("wilayah_id", "")
    date_from  = request.args.get("date_from", "")
    date_to    = request.args.get("date_to", "")

    where, params = ["1=1"], []
    if search:
        where.append("(e.container_id LIKE ? OR e.mac_address LIKE ? OR e.gateway_id LIKE ?)")
        params += [f"%{search}%"] * 3
    if event_type:
        where.append("e.event_type=?"); params.append(event_type)
    if depo_id:
        where.append("g.depo_id=?"); params.append(depo_id)
    elif wilayah_id:
        where.append("d.wilayah_id=?"); params.append(wilayah_id)
    if date_from:
        where.append("date(e.timestamp) >= ?"); params.append(date_from)
    if date_to:
        where.append("date(e.timestamp) <= ?"); params.append(date_to)

    where_sql = " AND ".join(where)
    db    = get_db()
    total = db.execute(f"""
        SELECT COUNT(*) FROM event_logs e
        LEFT JOIN gateways g ON e.gateway_id = g.gateway_id
        LEFT JOIN depo d ON g.depo_id = d.id
        WHERE {where_sql}
    """, params).fetchone()[0]

    rows = db.execute(f"""
        SELECT e.*, g.nama as gateway_nama, d.nama as depo_nama, w.nama as wilayah_nama
        FROM event_logs e
        LEFT JOIN gateways g ON e.gateway_id = g.gateway_id
        LEFT JOIN depo d ON g.depo_id = d.id
        LEFT JOIN wilayah w ON d.wilayah_id = w.id
        WHERE {where_sql}
        ORDER BY e.timestamp DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, (page - 1) * per_page]).fetchall()
    db.close()

    return jsonify({
        "total": total, "page": page, "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
        "data": [dict(r) for r in rows]
    })


# ─── API: Analytics ──────────────────────────────────────────────────────────

@app.route("/api/analytics/chart")
def api_analytics_chart():
    period     = request.args.get("period", "daily")
    depo_id    = request.args.get("depo_id", "")
    wilayah_id = request.args.get("wilayah_id", "")

    if period == "monthly":
        group_fmt, date_filter = "%Y-%m", "date(e.timestamp) >= date('now', '-12 months')"
    elif period == "yearly":
        group_fmt, date_filter = "%Y", "date(e.timestamp) >= date('now', '-5 years')"
    else:
        group_fmt, date_filter = "%Y-%m-%d", "date(e.timestamp) >= date('now', '-30 days')"

    gw_filter, params = "", []
    if depo_id:
        gw_filter = "AND g.depo_id = ?"; params = [depo_id]
    elif wilayah_id:
        gw_filter = "AND d.wilayah_id = ?"; params = [wilayah_id]

    db   = get_db()
    rows = db.execute(f"""
        SELECT strftime('{group_fmt}', e.timestamp) as period,
               e.event_type, COUNT(*) as cnt
        FROM event_logs e
        LEFT JOIN gateways g ON e.gateway_id = g.gateway_id
        LEFT JOIN depo d ON g.depo_id = d.id
        WHERE {date_filter} {gw_filter}
        GROUP BY period, e.event_type ORDER BY period
    """, params).fetchall()
    db.close()

    data = {}
    for r in rows:
        p = r["period"]
        if p not in data: data[p] = {"IN": 0, "OUT": 0}
        data[p][r["event_type"]] = r["cnt"]

    labels = sorted(data.keys())
    return jsonify({"labels": labels, "in": [data[l]["IN"] for l in labels], "out": [data[l]["OUT"] for l in labels]})


@app.route("/api/analytics/top_depo")
def api_analytics_top_depo():
    db   = get_db()
    rows = db.execute("""
        SELECT d.nama, COUNT(*) as total FROM event_logs e
        JOIN gateways g ON e.gateway_id = g.gateway_id
        JOIN depo d ON g.depo_id = d.id
        WHERE date(e.timestamp) >= date('now', '-30 days')
        GROUP BY d.nama ORDER BY total DESC LIMIT 5
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


# ─── API: Export ──────────────────────────────────────────────────────────────

@app.route("/api/export/excel")
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    where, params = _export_filters()
    db   = get_db()
    rows = db.execute(f"""
        SELECT e.timestamp, e.container_id, e.mac_address, e.gateway_id,
               g.nama as gateway_nama, d.nama as depo_nama, w.nama as wilayah_nama,
               e.event_type, e.rssi
        FROM event_logs e
        LEFT JOIN gateways g ON e.gateway_id = g.gateway_id
        LEFT JOIN depo d ON g.depo_id = d.id
        LEFT JOIN wilayah w ON d.wilayah_id = w.id
        WHERE {where} ORDER BY e.timestamp DESC
    """, params).fetchall()
    db.close()

    wb = Workbook(); ws = wb.active; ws.title = "Log IN/OUT Kontainer"
    headers = ["Timestamp","Container ID","BLE MAC","Gateway ID","Nama Gateway","Depo","Wilayah","Event","RSSI"]
    hfill = PatternFill("solid", fgColor="1a4a7a")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center")
    for ri, r in enumerate(rows, 2):
        vals = [r["timestamp"],r["container_id"],r["mac_address"],r["gateway_id"],
                r["gateway_nama"],r["depo_nama"],r["wilayah_nama"],r["event_type"],r["rssi"]]
        fc = "e8f4e8" if r["event_type"] == "IN" else "fde8e8"
        rf = PatternFill("solid", fgColor=fc)
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col, value=v)
            cell.fill = rf
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 4, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"log_inout_{datetime.date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/export/pdf")
def api_export_pdf():
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    where, params = _export_filters()
    db   = get_db()
    rows = db.execute(f"""
        SELECT e.timestamp, e.container_id, e.gateway_id,
               d.nama as depo_nama, e.event_type
        FROM event_logs e
        LEFT JOIN gateways g ON e.gateway_id = g.gateway_id
        LEFT JOIN depo d ON g.depo_id = d.id
        LEFT JOIN wilayah w ON d.wilayah_id = w.id
        WHERE {where} ORDER BY e.timestamp DESC LIMIT 500
    """, params).fetchall()
    db.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet(); story = []
    story.append(Paragraph("Laporan Log IN/OUT Kontainer", styles["Title"]))
    story.append(Paragraph(f"Dicetak: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 12))
    td = [["Timestamp","Container ID","Gateway","Depo","Event"]]
    for r in rows: td.append([r["timestamp"],r["container_id"],r["gateway_id"],r["depo_nama"] or "—",r["event_type"]])
    t = Table(td, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1a4a7a")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f5f5f5")]),
        ("GRID",(0,0),(-1,-1),.5,colors.grey),
    ]))
    story.append(t); doc.build(story); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"log_inout_{datetime.date.today().isoformat()}.pdf",
                     mimetype="application/pdf")


def _export_filters():
    where, params = ["1=1"], []
    if request.args.get("depo_id"):
        where.append("g.depo_id=?"); params.append(request.args["depo_id"])
    elif request.args.get("wilayah_id"):
        where.append("d.wilayah_id=?"); params.append(request.args["wilayah_id"])
    if request.args.get("date_from"):
        where.append("date(e.timestamp) >= ?"); params.append(request.args["date_from"])
    if request.args.get("date_to"):
        where.append("date(e.timestamp) <= ?"); params.append(request.args["date_to"])
    return " AND ".join(where), params


# ─── API: BLE Event (HTTP fallback) ──────────────────────────────────────────

@app.route("/api/ble/event", methods=["POST"])
def api_ble_event():
    """HTTP fallback — digunakan jika MQTT tidak tersedia atau untuk testing."""
    data       = request.json or {}
    mac        = data.get("mac_address", "")
    gateway_id = data.get("gateway_id", "")
    rssi       = int(data.get("rssi", -70))
    result     = process_ble_event(mac, gateway_id, rssi)
    status     = 200 if result["ok"] else 200  # selalu 200, biar Pi tidak retry
    return jsonify(result), status


# ─── API: Simulate ───────────────────────────────────────────────────────────

@app.route("/api/simulate", methods=["POST"])
def api_simulate():
    db       = get_db()
    tags     = db.execute("SELECT * FROM ble_tags WHERE status='Active'").fetchall()
    gateways = db.execute("SELECT * FROM gateways WHERE status='Online'").fetchall()
    db.close()
    if not tags or not gateways:
        return jsonify({"ok": False, "reason": "no active tags or online gateways"})
    tag    = random.choice(tags)
    gw     = random.choice(gateways)
    rssi   = random.randint(-85, -55)
    result = process_ble_event(tag["mac_address"], gw["gateway_id"], rssi)
    return jsonify({"ok": True, "tag": tag["mac_address"], "gateway": gw["gateway_id"], "result": result})


# ─── WebSocket ────────────────────────────────────────────────────────────────

@socketio.on("connect")
def on_connect():
    emit("connected", {"msg": "BLE Monitor connected"})


# ─── Gateway Heartbeat Checker ────────────────────────────────────────────────

def heartbeat_checker():
    """Set gateway Offline jika tidak ada aktivitas > 5 menit."""
    while True:
        import eventlet
        eventlet.sleep(60)
        db = get_db()
        db.execute("""
            UPDATE gateways SET status='Offline'
            WHERE last_active < datetime('now', '-5 minutes') AND status='Online'
        """)
        db.commit()
        db.close()


# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()
    socketio.start_background_task(heartbeat_checker)
    if MQTT_ENABLED:
        socketio.start_background_task(start_mqtt_client)
        log.info(f"MQTT subscriber dimulai → broker: {MQTT_HOST}:{MQTT_PORT}")
    else:
        log.info("MQTT dinonaktifkan (MQTT_ENABLED=false). Gunakan HTTP endpoint.")
    socketio.run(app, host="0.0.0.0", port=5000, debug=True, use_reloader=False)
